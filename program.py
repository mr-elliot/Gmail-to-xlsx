import imaplib, email, os, time, re
from email.header import decode_header
import dateutil.parser as dparser  # python-dateutil
from openpyxl import Workbook

start = time.time()  # for execution time
username = "YOUR GMAIL/G-SUITE ID"
password = "YOUR PASSWORD"

# Creating initial xlsx file for later purpose
wb = Workbook()
ws = wb.active
ws["A1"] = "Sender Name"
ws['B1'] = "Sender Email Id"
ws['C1'] = "Sent Date"
ws['D1'] = "Subject"
wb.save(filename="Dummy.xlsx")
r, c = 2, 1  # Default row and column

# Message's to be read from mail
N = 5

imap = imaplib.IMAP4_SSL("imap.gmail.com")
imap.login(username, password)
status, messages = imap.select("INBOX")

# total number of emails are in your account
messages = int(messages[0])


def link_remover(text):
    text = re.sub(r'https?://\S+', '', text)  # Link remover
    text = re.sub("\s\s+", " ", text)  # Multiple whitespace
    text = re.sub("\n\n+", "\n", text)  # Removing multiple lines
    return text


def info_extractor(fromm, da):
    mail = str(re.findall(r'[\w\.-]+@[\w\.-]+', fromm))  # Sender mail extractor
    mail = mail.translate(str.maketrans({"[": "", "]": "", "'": ""}))  # just to remove few character

    try:
        till = int(re.search("<", fromm).start())
        # till, _ = till.span()
        name = fromm[:till-1]  # Sender name extractor
    except AttributeError:
        # This error raise's when it can't find "<" in "fromm"
        name = fromm
    datee = str(dparser.parse(da))
    datee = datee[:11]  # Sender date extractor
    return name, mail, datee


def excel_work(name, mail_id, date, subject, r, c):
    items = [name, mail_id, date, subject]
    for item in items:
        ws.cell(row=r, column=c).value = item
        c += 1


for i in range(messages, messages-N, -1):
    res, msg = imap.fetch(str(i), "(RFC822)")
    for response in msg:
        if isinstance(response, tuple):
            msg = email.message_from_bytes(response[1])
            subject = decode_header(msg["Subject"])[0][0]
            if isinstance(subject, bytes):
                subject = subject.decode()
            from_ = msg.get("From")
            date_ = msg.get("Date")
            name, mail, datee = info_extractor(from_,date_)
            excel_work(name, mail, datee, subject, r, c)
            print("Subject:", subject)  # comment out if u dont need
            # if the email message is multipart
            if msg.is_multipart():
                # iterate over email parts
                for part in msg.walk():
                    content_type = part.get_content_type()
                    content_disposition = str(part.get("Content-Disposition"))
                    try:
                        body = part.get_payload(decode=True).decode()
                    except:
                        pass
                    if content_type == "text/plain" and "attachment" not in content_disposition:
                        if len(body) >= 1: print(link_remover(body))
            else:
                content_type = msg.get_content_type()
                body = msg.get_payload(decode=True).decode()
                if content_type == "text/plain":
                    if len(body) >= 1: print(link_remover(body))

            if content_type == "text/html":
                if not os.path.isdir("html files"):
                    # Inside the folder there will be your html files under subject name
                    os.mkdir("html files")
                filename = f"{subject[:50]}.html"
                filepath = os.path.join("html files", filename)
                try:
                    open(filepath, "w").write(body)  # write the file
                except FileNotFoundError:
                    print("FileNotFoundError")
                # webbrowser.open(filepath)  # Uncomment if u want to open every html file in web-browser(Not recommended since it will be annoying)
            print("="*100)
    r += 1
    c = 1
wb.save(filename="Converted.xlsx")
imap.close()
imap.logout()
end = time.time()
print("EXECUTION TIME IS: ", end - start)


